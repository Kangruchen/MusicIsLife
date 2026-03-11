"""Convert an Excel sheet named "节奏游戏对话表" to a JSON file suitable for Godot GDScript.

Requirements implemented:
 - Read Excel (openpyxl) sheet "节奏游戏对话表" and export dialog rows to JSON
 - Excel fields (expected header names):
     对话ID (int), 角色类型 (str), 角色名 (str), 对话文本 (str),
     立绘资源路径 (str, keep as string; if empty -> "null"),
     下一条ID (int), 触发场景 (str), 语音ID (str)
 - Export JSON that Godot can parse: Chinese not escaped, indent=4
 - Handle empty values (artwork empty -> string "null")
 - Print success message including number of dialogs loaded

Usage:
    python excel_to_godot_json.py input.xlsx output.json

Author: generated
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def _cell_to_str(cell_value: Any) -> str:
    """Convert a cell value to string; return empty string for None.

    Keeps existing strings as-is.
    """
    if cell_value is None:
        return ""
    return str(cell_value)


def convert_excel_to_godot_json(
    excel_path: str | Path,
    json_path: str | Path,
    sheet_name: str = "节奏游戏对话表",
) -> int:
    """Read the Excel file and export the dialog table to JSON.

    Returns the number of dialog entries written.

    Data shape for each entry in the output list:
      {
        "id": int,
        "role_type": str,
        "role_name": str,
        "text": str,
        "artwork": str,        # string "null" when the cell is empty
        "next_id": Optional[int],
        "trigger_scene": str,
        "voice_id": str,
      }

    Notes / assumptions:
    - The sheet must contain a header row with the exact column names
      (whitespace tolerant):
        对话ID, 角色类型, 角色名, 对话文本, 立绘资源路径, 下一条ID, 触发场景, 语音ID
    - Rows with empty 对话ID are skipped.
    - 下一条ID when empty becomes null in JSON.
    - 立绘资源路径 when empty becomes the LITERAL string "null" (per requirement).
    """
    excel_path = Path(excel_path)
    # Allow Godot-style res:// paths by resolving them to the project root.
    # We assume this script is inside the project and project root is two parents up
    # (playground/dialogue_test -> playground -> project root).
    json_path_raw = str(json_path)
    if json_path_raw.startswith("res://"):
        # map res://some/path.json -> <repo_root>/some/path.json
        repo_root = Path(__file__).resolve().parents[2]
        rel = json_path_raw[len("res://"):].lstrip("/\\")
        json_path = repo_root.joinpath(rel)
    else:
        json_path = Path(json_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(filename=str(excel_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Read header row (first non-empty row). We assume header is the first row.
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    # Map normalized header -> column index
    header_map: Dict[str, int] = {}
    for idx, raw in enumerate(header_cells):
        if raw is None:
            continue
        key = str(raw).strip().replace("\uFEFF", "")  # remove BOM if present
        header_map[key] = idx

    # Expected headers (in Chinese) mapped to internal keys
    expected_headers = {
        "对话ID": "id",
        "角色类型": "role_type",
        "角色名": "role_name",
        "对话文本": "text",
        "立绘资源路径": "artwork",
        "下一条ID": "next_id",
        "触发场景": "trigger_scene",
        "语音ID": "voice_id",
    }

    # Verify required headers exist
    missing = [h for h in expected_headers.keys() if h not in header_map]
    if missing:
        raise ValueError(f"Missing required headers in sheet '{sheet_name}': {missing}")

    dialogs: List[Dict[str, Any]] = []

    # Iterate rows after header
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Read by header indices
        raw_id = row[header_map["对话ID"]]
        if raw_id is None:
            # skip rows without an ID
            continue

        try:
            dialog_id = int(raw_id)
        except Exception:
            # If ID isn't a number, try to coerce from string
            dialog_id = int(str(raw_id).strip())

        # Strings
        role_type = _cell_to_str(row[header_map["角色类型"]])
        role_name = _cell_to_str(row[header_map["角色名"]])
        text = _cell_to_str(row[header_map["对话文本"]])

        # Artwork: if empty -> string "null" (literal), otherwise string value
        raw_artwork = row[header_map["立绘资源路径"]]
        if raw_artwork is None or (isinstance(raw_artwork, str) and raw_artwork.strip() == ""):
            artwork = "null"
        else:
            artwork = str(raw_artwork)

        # next_id: allow empty -> None (will become JSON null), otherwise int
        raw_next = row[header_map["下一条ID"]]
        if raw_next is None or (isinstance(raw_next, str) and str(raw_next).strip() == ""):
            next_id: Optional[int] = None
        else:
            try:
                next_id = int(raw_next)
            except Exception:
                # try to parse numeric string
                next_id = int(str(raw_next).strip())

        trigger_scene = _cell_to_str(row[header_map["触发场景"]])
        voice_id = _cell_to_str(row[header_map["语音ID"]])

        entry: Dict[str, Any] = {
            "id": dialog_id,
            "role_type": role_type,
            "role_name": role_name,
            "text": text,
            "artwork": artwork,
            "next_id": next_id,
            "trigger_scene": trigger_scene,
            "voice_id": voice_id,
        }

        dialogs.append(entry)

    # Write JSON with Chinese characters preserved and indent of 4 spaces
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(dialogs, f, ensure_ascii=False, indent=4)

    # Print success message
    print(f"Successfully exported {len(dialogs)} dialogs to '{json_path}'")
    return len(dialogs)


def _cli_main() -> None:
    parser = argparse.ArgumentParser(description="Convert Excel dialog sheet to Godot-friendly JSON")
    parser.add_argument("excel", help="Path to the input Excel file (xlsx)")
    parser.add_argument("json", help="Path to the output JSON file")
    parser.add_argument("--sheet", default="节奏游戏对话表", help="Sheet name to read from")
    args = parser.parse_args()

    try:
        count = convert_excel_to_godot_json(args.excel, args.json, sheet_name=args.sheet)
    except Exception as e:
        print(f"Error: {e}")
        raise SystemExit(1)

    print(f"Loaded dialogs: {count}")


if __name__ == "__main__":
    _cli_main()


# Example function call (uncomment to run inside a script):
# from pathlib import Path
# excel_file = Path(r"./dialogues.xlsx")
# # Export to Godot-style path. This will write to the project root `data/dialogue_data.json`.
# convert_excel_to_godot_json(excel_file, "res://data/dialogue_data.json")
